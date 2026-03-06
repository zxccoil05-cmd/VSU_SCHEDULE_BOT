import openpyxl
import requests
from io import BytesIO
import logging
from bs4 import BeautifulSoup
from urllib.parse import urljoin

logger = logging.getLogger(__name__)

class HumParser:
    def __init__(self):
        # Актуальная ссылка на Гуманитарный факультет (ФГЗиК)
        self.base_page_url = "https://vsu.by/universitet/fakultety/fakultet-gumanitarnogo-znaniya-i-kommunikacij/raspisanie.html"
        self.cache = {}

    def _get_value(self, sheet, row, col):
        """Пробивает объединенные ячейки"""
        for merged in sheet.merged_cells.ranges:
            if row in range(merged.min_row, merged.max_row + 1) and \
               col in range(merged.min_col, merged.max_col + 1):
                return sheet.cell(row=merged.min_row, column=merged.min_col).value
        return sheet.cell(row=row, column=col).value

    async def _find_all_links(self):
        """Более гибкий поиск ссылок для ГФ"""
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(self.base_page_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Ищем все ссылки в основном контенте страницы
            links = soup.find_all('a', href=True)
            
            found_urls = []
            # Убираем только то, что 100% не является дневным расписанием
            black_list = ["заоч", "зфо", "зфпо", "зачет", "экзамен", "сессия", "магистр", "практик", "экз_"]

            for l in links:
                text = l.get_text(separator=" ", strip=True).lower()
                href = l['href'].lower()
                
                # Проверяем, что это Excel
                if ".xlsx" in href or ".xls" in href:
                    # Если в тексте есть 'расписание' и НЕТ слов из черного списка
                    if "расписание" in text or "курс" in text:
                        if not any(word in text for word in black_list) and \
                           not any(word in href for word in black_list):
                            full_url = urljoin(self.base_page_url, l['href'])
                            found_urls.append(full_url)
            
            # Если ничего не нашли, попробуем взять вообще все xlsx без черного списка
            if not found_urls:
                for l in links:
                    href = l['href'].lower()
                    if (".xlsx" in href or ".xls" in href) and not any(w in href for w in black_list):
                        found_urls.append(urljoin(self.base_page_url, l['href']))

            return list(set(found_urls))
        except Exception as e:
            logger.error(f"Ошибка поиска ссылок ГФ: {e}")
            return []

    async def refresh(self):
        """Парсинг ГФ с объединением подгрупп и фильтром названий"""
        links = await self._find_all_links()
        if not links:
            logger.error("❌ Ссылки для ГФ не найдены")
            return None

        all_fac_data = {}
        
        for url in links:
            try:
                logger.info(f"📥 Обработка файла ГФ: {url}")
                response = requests.get(url, timeout=20)
                wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
                sheet = wb.active
                
                subgroups = {}
                group_row = 14
                
                # ШАГ 1: Поиск номеров групп (строки 12-16)
                for r_check in [13, 14, 15, 12, 16]:
                    for col in range(4, sheet.max_column + 1):
                        val = sheet.cell(row=r_check, column=col).value
                        if val:
                            val_str = str(val).strip().replace('\n', ' ')
                            # Код группы: короткий (<15 симв) и содержит цифры
                            if 2 < len(val_str) < 15 and any(c.isdigit() for c in val_str):
                                # Проверяем ячейку ниже на подгруппу (1 или 2)
                                sub_val = sheet.cell(row=r_check + 1, column=col).value
                                if sub_val and str(sub_val).strip() in ["1", "2"]:
                                    val_str = f"{val_str}_{str(sub_val).strip()}"
                                
                                subgroups[col] = val_str
                    
                    if subgroups:
                        group_row = r_check
                        # Если нашли подгруппы строкой ниже, смещаем начало данных
                        if any("_" in name for name in subgroups.values()):
                            group_row += 1
                        break

                if not subgroups: continue

                # ШАГ 2: Сбор данных
                for col_idx, g_name in subgroups.items():
                    if g_name not in all_fac_data:
                        all_fac_data[g_name] = {}
                    
                    current_day = "Неизвестно"
                    r = group_row + 1
                    while r < 350:
                        day_val = str(self._get_value(sheet, r, 1) or "").strip()
                        if not day_val:
                            day_val = str(self._get_value(sheet, r, 2) or "").strip()
                        
                        if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                            current_day = day_val
                            if current_day not in all_fac_data[g_name]:
                                all_fac_data[g_name][current_day] = []

                        pair_num = str(self._get_value(sheet, r, 3) or "").strip()
                        if pair_num.isdigit():
                            subject = self._get_value(sheet, r, col_idx)
                            if subject:
                                time_raw = str(self._get_value(sheet, r + 1, 3) or "")
                                teacher = self._get_value(sheet, r + 1, col_idx)
                                room = self._get_value(sheet, r + 2, col_idx)
                                
                                entry = {
                                    "time": time_raw.replace("(", "").replace(")", "").strip(),
                                    "name": str(subject).strip().replace('\n', ' '),
                                    "teacher": str(teacher).strip() if teacher else "---",
                                    "room": str(room).strip() if room else "---"
                                }
                                # Проверка на дубликаты (для потоковых лекций)
                                if entry not in all_fac_data[g_name][current_day]:
                                    all_fac_data[g_name][current_day].append(entry)
                            r += 3
                        else:
                            r += 1
            except Exception as e:
                logger.error(f"⚠️ Ошибка ГФ {url}: {e}")

        self.cache = all_fac_data
        logger.info(f"✅ ГФ обновлен. Всего групп: {len(all_fac_data)}")
        return all_fac_data

    def get_groups(self):
        return sorted(list(self.cache.keys())) if self.cache else []

    def get_schedule(self, group_name):
        return self.cache.get(group_name, {}) if self.cache else {}