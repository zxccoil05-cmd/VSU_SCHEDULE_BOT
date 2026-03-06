import openpyxl
import requests
from io import BytesIO
import logging
from bs4 import BeautifulSoup
from urllib.parse import urljoin

logger = logging.getLogger(__name__)

class FSPIPParser:
    def __init__(self):
        # URL страницы расписания ФСПиП
        self.base_page_url = "https://vsu.by/universitet/fakultety/sotsialnoj-pedagogiki-i-psikhologii/raspisanie.html"
        self.cache = {}

    def _get_value(self, sheet, row, col):
        """Пробивает объединенные ячейки"""
        for merged in sheet.merged_cells.ranges:
            if row in range(merged.min_row, merged.max_row + 1) and \
               col in range(merged.min_col, merged.max_col + 1):
                return sheet.cell(row=merged.min_row, column=merged.min_col).value
        return sheet.cell(row=row, column=col).value

    async def _find_all_links(self):
        """Ищет только актуальные файлы ДФПО, игнорируя мусор"""
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(self.base_page_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            links = soup.find_all('a', href=True)
            
            found_urls = []
            # Жесткий фильтр: только ДФПО, без зачетов и заочки
            black_list = ["заоч", "зфо", "зфпо", "зачет", "экзамен", "сессия", "магистр", "зач_", "экз_"]

            for l in links:
                text = l.get_text(separator=" ", strip=True).lower()
                href = l['href'].lower()
                
                if (".xlsx" in href or ".xls" in href) and "расписание" in text:
                    # Оставляем только дневную форму (ДФПО)
                    if "дфпо" in text or "дфпо" in href:
                        if not any(word in text for word in black_list) and \
                           not any(word in href for word in black_list):
                            full_url = urljoin(self.base_page_url, l['href'])
                            found_urls.append(full_url)
            
            return list(set(found_urls))
        except Exception as e:
            logger.error(f"Ошибка поиска ссылок ФСПиП: {e}")
            return []

    async def refresh(self):
        """Парсинг файлов с объединением подгрупп и очисткой названий групп"""
        links = await self._find_all_links()
        if not links:
            logger.error("❌ Ссылки для ФСПиП не найдены")
            return None

        all_fac_data = {}
        
        for url in links:
            try:
                logger.info(f"📥 Обработка файла ФСПиП: {url}")
                response = requests.get(url, timeout=20)
                wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
                sheet = wb.active
                
                subgroups = {}
                group_row = 14
                
                # Поиск реальных номеров групп (строки 13-16)
                for r_check in [14, 15, 13, 16]:
                    for col in range(4, sheet.max_column + 1):
                        val = sheet.cell(row=r_check, column=col).value
                        if val:
                            val_str = str(val).strip().replace('\n', ' ')
                            # Код группы: короткий и содержит цифры
                            if 2 < len(val_str) < 15 and any(c.isdigit() for c in val_str):
                                # Проверка на подгруппу строкой ниже
                                sub_val = sheet.cell(row=r_check + 1, column=col).value
                                if sub_val and str(sub_val).strip() in ["1", "2"]:
                                    val_str = f"{val_str}_{str(sub_val).strip()}"
                                
                                subgroups[col] = val_str
                    
                    if subgroups:
                        group_row = r_check
                        if "_" in list(subgroups.values())[0]:
                            group_row += 1
                        break

                if not subgroups: continue

                for col_idx, g_name in subgroups.items():
                    if g_name not in all_fac_data:
                        all_fac_data[g_name] = {}
                    
                    current_day = "Неизвестно"
                    r = group_row + 1
                    while r < 350:
                        day_val = str(self._get_value(sheet, r, 1) or "").strip()
                        if not day_val:
                            day_val = str(self._get_value(sheet, r, 2) or "").strip()
                        
                        if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                            current_day = day_val
                            if current_day not in all_fac_data[g_name]:
                                all_fac_data[g_name][current_day] = []

                        pair_num = str(self._get_value(sheet, r, 3) or "").strip()
                        if pair_num.isdigit():
                            subject = self._get_value(sheet, r, col_idx)
                            if subject:
                                time_raw = str(self._get_value(sheet, r + 1, 3) or "")
                                teacher = self._get_value(sheet, r + 1, col_idx)
                                room = self._get_value(sheet, r + 2, col_idx)
                                
                                entry = {
                                    "time": time_raw.replace("(", "").replace(")", "").strip(),
                                    "name": str(subject).strip().replace('\n', ' '),
                                    "teacher": str(teacher).strip() if teacher else "---",
                                    "room": str(room).strip() if room else "---"
                                }
                                # Проверка на дубликаты (для общих лекций)
                                if entry not in all_fac_data[g_name][current_day]:
                                    all_fac_data[g_name][current_day].append(entry)
                            r += 3
                        else:
                            r += 1
            except Exception as e:
                logger.error(f"⚠️ Ошибка ФСПиП {url}: {e}")

        self.cache = all_fac_data
        logger.info(f"✅ ФСПиП обновлен. Групп: {len(all_fac_data)}")
        return all_fac_data

    def get_groups(self):
        return sorted(list(self.cache.keys())) if self.cache else []

    def get_schedule(self, group_name):
        return self.cache.get(group_name, {}) if self.cache else {}