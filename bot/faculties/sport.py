import openpyxl
import requests
from io import BytesIO
import logging
from bs4 import BeautifulSoup
from urllib.parse import urljoin

logger = logging.getLogger(__name__)

class SportParser:
    def __init__(self):
        # URL страницы расписания ФФКиС
        self.base_page_url = "https://vsu.by/universitet/fakultety/fizicheskoj-kultury-i-sporta/raspisanie.html"
        self.cache = None

    def _get_value(self, sheet, row, col):
        """Твоя оригинальная функция для пробития объединенных ячеек"""
        for merged in sheet.merged_cells.ranges:
            if row in range(merged.min_row, merged.max_row + 1) and \
               col in range(merged.min_col, merged.max_col + 1):
                return sheet.cell(row=merged.min_row, column=merged.min_col).value
        return sheet.cell(row=row, column=col).value

    async def _find_actual_link(self):
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(self.base_page_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            links = soup.find_all('a', href=True)
            
            for l in links:
                link_text = l.get_text(separator=" ", strip=True)
                link_href = l['href']
                # Ищем "Расписание", "занятий" и ".xlsx"
                if "Расписание" in link_text and "занятий" in link_text and (".xlsx" in link_href or ".xls" in link_href):
                    if "заоч" not in link_text.lower():
                        return urljoin(self.base_page_url, link_href)
            return None
        except Exception as e:
            logger.error(f"Ошибка поиска ссылки ФФКиС: {e}")
            return None

    async def refresh(self):
        actual_url = await self._find_actual_link()
        if not actual_url: return None

        try:
            logger.info(f"Скачивание файла ФФКиС: {actual_url}")
            response = requests.get(actual_url)
            wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
            sheet = wb.active
            
            subgroups = {}
            # Пробуем найти группы в 14-й строке
            for col in range(4, sheet.max_column + 1):
                name = sheet.cell(row=14, column=col).value
                if name and str(name).strip() and len(str(name).strip()) > 1:
                    subgroups[col] = str(name).strip()
            
            final_data = {}
            for col_idx, sg_name in subgroups.items():
                final_data[sg_name] = {}
                current_day = "Неизвестно"
                r = 16
                while r < 400: # У спортсменов бывает много пар
                    day_val = str(self._get_value(sheet, r, 1) or "").strip()
                    if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                        current_day = day_val
                        if current_day not in final_data[sg_name]:
                            final_data[sg_name][current_day] = []

                    pair_num = str(self._get_value(sheet, r, 3) or "").strip()
                    if pair_num.isdigit():
                        subject = self._get_value(sheet, r, col_idx)
                        time_raw = str(self._get_value(sheet, r + 1, 3) or "")
                        teacher = self._get_value(sheet, r + 1, col_idx)
                        room = self._get_value(sheet, r + 2, col_idx)
                        
                        if subject:
                            final_data[sg_name][current_day].append({
                                "time": time_raw.replace("(", "").replace(")", "").strip(),
                                "name": str(subject).strip().replace('\n', ' '),
                                "teacher": str(teacher).strip() if teacher else "---",
                                "room": str(room).strip() if room else "---"
                            })
                        r += 3
                    else:
                        r += 1
            
            self.cache = final_data
            logger.info("✅ ФФКиС успешно обновлен")
            return final_data
        except Exception as e:
            logger.error(f"Ошибка парсинга Excel ФФКиС: {e}")
            return None

    def get_groups(self):
        return list(self.cache.keys()) if self.cache else []

    def get_schedule(self, group_name):
        return self.cache.get(group_name, {}) if self.cache else {}
