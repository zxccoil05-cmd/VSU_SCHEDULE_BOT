import openpyxl
import requests
from io import BytesIO
import logging
from bs4 import BeautifulSoup
from urllib.parse import urljoin

logger = logging.getLogger(__name__)

class PedParser:
    def __init__(self):
        # Актуальная ссылка на Педфак
        self.base_page_url = "https://vsu.by/universitet/fakultety/pedagogicheskij-fakultet/raspisanie.html"
        self.cache = {}

    def _get_value(self, sheet, row, col):
        """Пробивает объединенные ячейки"""
        for merged in sheet.merged_cells.ranges:
            if row in range(merged.min_row, merged.max_row + 1) and \
               col in range(merged.min_col, merged.max_col + 1):
                return sheet.cell(row=merged.min_row, column=merged.min_col).value
        return sheet.cell(row=row, column=col).value

    async def _find_all_links(self):
        """Ищем только файлы ДФПО (дневное), игнорируя мусор"""
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(self.base_page_url, headers=headers, timeout=15)
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            links = soup.find_all('a', href=True)
            
            found_urls = []
            # Стоп-слова: убираем заочку, зачеты и прочее
            black_list = ["заоч", "зфо", "зфпо", "зачет", "экзамен", "сессия", "магистр", "экз_", "зач_"]

            for l in links:
                text = l.get_text(separator=" ", strip=True).lower()
                href = l['href'].lower()
                
                # Ищем файлы Excel, которые относятся к расписанию
                if (".xlsx" in href or ".xls" in href) and ("расписание" in text or "курс" in text):
                    # Проверка на отсутствие мусора
                    if not any(word in text for word in black_list) and \
                       not any(word in href for word in black_list):
                        full_url = urljoin(self.base_page_url, l['href'])
                        found_urls.append(full_url)
            
            unique_urls = list(set(found_urls))
            logger.info(f"🔎 Найдено файлов ПФ: {len(unique_urls)}")
            return unique_urls
        except Exception as e:
            logger.error(f"Ошибка поиска ссылок ПФ: {e}")
            return []

    async def refresh(self):
        """Парсинг ПФ с умной фильтрацией названий групп и подгрупп"""
        links = await self._find_all_links()
        if not links:
            logger.error("❌ Ссылки для ПФ не найдены")
            return None

        all_fac_data = {}
        
        for url in links:
            try:
                logger.info(f"📥 Обработка файла ПФ: {url}")
                response = requests.get(url, timeout=20)
                wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
                sheet = wb.active
                
                subgroups = {}
                group_row = 14
                
                # ШАГ 1: Поиск номеров групп (строки 12-16)
                # Условие len < 15 отсекает длинные названия специальностей
                for r_check in [14, 15, 13, 16, 12]:
                    for col in range(4, sheet.max_column + 1):
                        val = sheet.cell(row=r_check, column=col).value
                        if val:
                            val_str = str(val).strip().replace('\n', ' ')
                            # Группа — это короткое название, содержащее цифру
                            if 2 < len(val_str) < 15 and any(c.isdigit() for c in val_str):
                                # Проверка на подгруппу строкой ниже (1 или 2)
                                sub_val = sheet.cell(row=r_check + 1, column=col).value
                                if sub_val and str(sub_val).strip() in ["1", "2"]:
                                    val_str = f"{val_str}_{str(sub_val).strip()}"
                                
                                subgroups[col] = val_str
                    
                    if subgroups:
                        group_row = r_check
                        # Если нашли подгруппы, пропускаем еще одну строку
                        if any("_" in name for name in subgroups.values()):
                            group_row += 1
                        break

                if not subgroups: continue

                # ШАГ 2: Сбор пар
                for col_idx, g_name in subgroups.items():
                    if g_name not in all_fac_data:
                        all_fac_data[g_name] = {}
                    
                    current_day = "Неизвестно"
                    r = group_row + 1
                    while r < 350:
                        # Поиск дня недели
                        day_val = str(self._get_value(sheet, r, 1) or "").strip()
                        if not day_val:
                            day_val = str(self._get_value(sheet, r, 2) or "").strip()
                        
                        if day_val in ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']:
                            current_day = day_val
                            if current_day not in all_fac_data[g_name]:
                                all_fac_data[g_name][current_day] = []

                        # Проверка номера пары
                        pair_num = str(self._get_value(sheet, r, 3) or "").strip()
                        if pair_num.isdigit():
                            subject = self._get_value(sheet, r, col_idx)
                            if subject:
                                time_raw = str(self._get_value(sheet, r + 1, 3) or "")
                                teacher = self._get_value(sheet, r + 1, col_idx)
                                room = self._get_value(sheet, r + 2, col_idx)
                                
                                entry = {
                                    "time": time_raw.replace("(", "").replace(")", "").strip(),
                                    "name": str(subject).strip().replace('\n', ' '),
                                    "teacher": str(teacher).strip() if teacher else "---",
                                    "room": str(room).strip() if room else "---"
                                }
                                if entry not in all_fac_data[g_name][current_day]:
                                    all_fac_data[g_name][current_day].append(entry)
                            r += 3
                        else:
                            r += 1
            except Exception as e:
                logger.error(f"⚠️ Ошибка ПФ {url}: {e}")

        self.cache = all_fac_data
        logger.info(f"✅ ПФ обновлен. Всего групп: {len(all_fac_data)}")
        return all_fac_data

    def get_groups(self):
        return sorted(list(self.cache.keys())) if self.cache else []

    def get_schedule(self, group_name):
        return self.cache.get(group_name, {}) if self.cache else {}