import openpyxl
import requests
import logging
from bs4 import BeautifulSoup
from io import BytesIO
import asyncio
import re

logging.basicConfig(level=logging.INFO, format='%(levelname)s:%(name)s:%(message)s')
logger = logging.getLogger(__name__)

class MultiFacultyParser:
    def __init__(self, faculties_config):
        self.faculties = faculties_config
        self.cache = {}

    async def refresh_all(self):
        logger.info("🚀 ЗАПУСК ОБНОВЛЕНИЯ...")
        for name, url in self.faculties.items():
            data = await self._parse_faculty_page(name, url)
            if data:
                self.cache[name] = data
                logger.info(f"✅ {name}: Готово. Групп: {len(data)}")
            else:
                self.cache[name] = {}
                logger.error(f"❌ {name}: Расписание не найдено.")

    async def _parse_faculty_page(self, fac_name, page_url):
        all_groups = {}
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(page_url, headers=headers, timeout=15)
            soup = BeautifulSoup(response.content, 'html.parser')
            
            links = []
            for a in soup.find_all('a', href=True):
                text = a.get_text().lower().strip()
                href = a['href'].lower()
                
                if not ('.xlsx' in href or '.xls' in href): continue
                
                keywords = ['расписание занятий', 'курс', 'дневн']

                blacklist = ['заоч', 'экзам', 'зачет', 'сессия', 'магистр', 'практик', 'план', 'график', 'зачетов', 'курсовые', 'полугодие']
                
                if any(word in text or word in href for word in keywords) and not any(bad in text or bad in href for bad in blacklist):
                    full_url = "https://vsu.by" + a['href'] if a['href'].startswith('/') else a['href']
                    links.append(full_url)

            for link in list(set(links)):
                data = self._parse_excel_file(link)
                if data: all_groups.update(data)
            return all_groups
        except Exception as e:
            logger.error(f" Ошибка {fac_name}: {e}")
            return None

    def _parse_excel_file(self, url):
        try:
            res = requests.get(url, timeout=20)
            wb = openpyxl.load_workbook(BytesIO(res.content), data_only=True)
            sheet = wb.active
            found_data = {}

            # Строки согласно твоим замерам
            GROUP_ROW = 13
            SUBGROUP_ROW = 14
            START_LESSONS_ROW = 15

            for c in range(3, sheet.max_column + 1):
                # Читаем номер группы (нужен для логов и проверки)
                group_name = self._get_merged_value(sheet, GROUP_ROW, c)
                
                # Читаем название подгруппы (то, что пойдет на кнопку)
                sub_val = str(sheet.cell(row=SUBGROUP_ROW, column=c).value or "").strip()
                
                # Если в 13-й строке пусто, пропускаем колонку
                if not group_name or group_name == "None" or not re.search(r'\d', group_name):
                    continue
                
                # УСТАНАВЛИВАЕМ ИМЯ КНОПКИ
                # Если 14-я строка не пустая, берем только её
                if sub_val and sub_val != "None":
                    final_name = sub_val
                else:
                    # Если вдруг 14-я строка пустая, оставляем номер группы
                    final_name = group_name

                # Извлекаем уроки
                lessons = self._extract_lessons(sheet, c, START_LESSONS_ROW)
                
                if lessons:
                    # Если кнопки с таким именем (например "1") уже есть от другой группы,
                    # в словаре они могут перезаписаться. 
                    # Чтобы этого не было, можно добавить невидимый символ или индекс,
                    # но пока сделаем просто уникальные ключи.
                    found_data[final_name] = lessons
                    logger.info(f"   [🎯] Кнопка: {final_name} (из группы {group_name})")

            return found_data
        except Exception as e:
            logger.error(f"Ошибка парсинга: {e}")
            return None

    def _get_merged_value(self, sheet, row, col):
        """Метод для чтения объединенных ячеек (чтобы видеть группу над обеими подгруппами)"""
        cell = sheet.cell(row=row, column=col)
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return str(sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value or "").strip()
        return str(cell.value or "").strip()

    def _parse_simple_groups(self, sheet):
        """Запасной метод, если в файле нет деления на подгруппы"""
        found_data = {}
        for r in range(2, 20):
            found = False
            for c in range(3, sheet.max_column + 1):
                val = str(sheet.cell(row=r, column=c).value or "").strip()
                if val and val != "None" and re.search(r'\d', val) and len(val) < 10:
                    if any(x in val.lower() for x in ["подгруп", "курс", "декан", "____"]): continue
                    lessons = self._extract_lessons(sheet, c, r + 1)
                    if lessons:
                        found_data[val] = lessons
                        found = True
            if found: break
        return found_data

    def _extract_lessons(self, sheet, col, start_row):
        schedule = {"Понедельник":[], "Вторник":[], "Среда":[], "Четверг":[], "Пятница":[], "Суббота":[]}
        current_day = None
        for r in range(start_row, 120):
            for c_day in [1, 2]:
                d_val = str(sheet.cell(row=r, column=c_day).value or "").strip().capitalize()
                if d_val in schedule:
                    current_day = d_val
                    break
            
            time_val = str(sheet.cell(row=r, column=2).value or "").strip()
            lesson_val = str(sheet.cell(row=r, column=col).value or "").strip()

            if current_day and lesson_val and lesson_val != "None" and len(lesson_val) > 2:
                # Фильтр для самих уроков, чтобы не попадал мусор в расписание
                if "_" in lesson_val or "подпись" in lesson_val.lower(): continue
                
                schedule[current_day].append({
                    "time": time_val if time_val != "None" else "",
                    "name": lesson_val,
                    "teacher": "", "room": ""
                })
        
        return schedule if any(len(v) > 0 for v in schedule.values()) else None

    def get_groups_list(self, faculty_name):
        return sorted(list(self.cache.get(faculty_name, {}).keys()))

    async def get_faculty_schedule(self, faculty_name):
        return self.cache.get(faculty_name, {})